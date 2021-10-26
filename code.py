## python version 3.7.4
import openpyxl as openxl
from spellchecker import SpellChecker
import urllib.parse
import urllib.request
import os
from bs4 import BeautifulSoup
import nltk

import re
from nltk.corpus import stopwords
from nltk.tokenize import word_tokenize
import numpy as np
import scipy as sp

from string import ascii_uppercase

import seaborn as sns   
import pandas as pd

## Problem 1 & 2

def read_keywords(excel_file, keywords, cell):
    # Reads Keywords from an excel(.xlsx) file
    workbook = openxl.load_workbook(excel_file)
    sheet = workbook.active
    spell = SpellChecker()
    while sheet["A{}".format(cell)].value != None:
        keywords.append(sheet["A{}".format(cell)].value)
        keywords[-1] = " ".join([spell.correction(word).capitalize() for word in keywords[-1].split()])
        # Ensures that keywords have correct spellings
        cell += 1
    workbook.close()
    # Keywords are returned as a list
    return keywords

def save_articles(keywords):
    # facilitates compiling search results, extracting article data and saving them to the file system
    for word in keywords:
        values = {'q' : word}
        data = urllib.parse.urlencode(values)
        search_url = "{}search?{}".format(url, data)
        # URL for searching a keyword using BBC's search feature

        articles = compile_articles(word, "", 0, search_url)
        # The string 'articles' will contain all data for 100 articles of a certain keyword
        print("Completed downloading relevant articles for Keyword {}\n".format(word))

        with open("data/articles/{}.txt".format(word),"w", encoding = "utf-8") as f:
            # Data is written to a text file using the name of the keyword
            f.write(articles) 
            f.close()

def compile_articles(keyword, articles, counter, search_url):
    print("Searching for articles on {}\n".format(keyword))
    search_soup = create_soup(search_url)
    # perform a search on each keyword using bbc's search feature

    num_pages = find_num_pages(search_soup)
    # Obtains the number of pages of search results returned on a BBC search query

    for page in range(1,num_pages+1):
        print("Searching Page {} for articles on {}".format(page,keyword))
        # Loops through each search result page
        if page != 1:
            search_soup = create_soup("{}&page={}".format(search_url, page))

        links = search_soup.find_all('ul')[3].find_all('a', href = True)
        # Obtains all article links found on a single search page

        for link in links:
            # Loops through each link found on a seach page
            if "www.bbc.co.uk/news" in link['href']:
                # All older BBC articles are rejected
                article = read_article(link['href'], "")
                # Retrieves all relevant article text 
                if article == None:
                    continue
                if keyword.lower() not in article:
                    # Ensures that the article contains the keyword
                    if ' ' not in keyword:
                        continue
                    else:
                        key_list, b = [keyword.lower()][0].split(), True
                        for word in key_list:
                            if word not in article:
                                b = False
                                break
                        if b == False:
                            continue

                articles += article
                # Validated article gets appended to the string "articles"
                counter += 1
            if counter == num_articles:
                # If 100 articles have been verified and stored, a string, 'articles' which contains all required data is returned
                return articles
        print("{} relevant articles found so far\n".format(counter))
    return articles

def create_soup(link):
    # Helper function, facilitatates creation of a soup from raw article data using BeautifulSoup
    request = urllib.request.Request(link)
    response = urllib.request.urlopen(request)
    return BeautifulSoup(response.read(), "html.parser")

def find_num_pages(soup):
    # Finds the number of pages from a search query on a keyword
    try:
        num_pages = int(soup.find_all('b')[1].get_text())
    except IndexError:
        num_pages = 1
    # Number of Pages is returned
    return num_pages

def read_article(article_link, article_text):
    # This function downloads the HTML from an article's webpage and then extracts the article's title and content
    article_soup = create_soup(article_link)
    # Article's webpage is downloaded(Problem 1)
    article_lines = [item.text.lower() for item in article_soup.find_all("div", attrs={"data-component" : "text-block"})]
    # Relevant content from the webpage is stored in the list 'article_lines'
    if article_lines == []:
        return None
    else:
        article_text += article_soup.find('title').text.replace(" - BBC News ",'.\n').lower()
        # Title of the article is extracted
        article_text += '\n'.join(article_lines)
        article_text += separator 
    return article_text # Article data returned as a String

## Problem 3

model = {}

def find_keyword_distances(keywords, data_dictionary, num_words):
    # facilitates cleaning of raw data and then feeding it to the semantic distance algorithm
    load_glove_embedding()
    # GloVe mode loaded

    semantic_distances = np.zeros((len(keywords),len(keywords)), dtype=float)
    # sematic_distances(2d array) will store the scores obtained by comparing article data of each keyword

    for word in keywords:
        # Reads article data from text files and convert them to an acceptable format
        with open("data/articles/{}.txt".format(word),"r", encoding = "utf-8") as f:
            full_dictionary = format_text(f.read(), {})
            shortened_dictionary = shorten_dictionary(full_dictionary, {}, num_words)
            data_dictionary[word] = shortened_dictionary
            print("Article data for keyword, {} has been loaded and cleaned".format(word))
            f.close()

    for i in range(0, len(keywords)):
        # Compares article data of every keyword and calculates the semantic distance between them
        for j in range(i, len(keywords)):
            distance = calculate_distance(data_dictionary[keywords[i]],data_dictionary[keywords[j]], 0, 0)
            # distance = distance between ith and jth keyword
            if i != j:
                semantic_distances[i][j], semantic_distances[j][i] = distance, distance
            else:
                semantic_distances[i][j] = distance
        print("Finished calculating distances for the keyword, {}".format(keywords[i]))

    distance_matrix = np.zeros((len(keywords),len(keywords)), dtype=float)
    # distance_matrix(2d array) will store the similarity between all keywords as a percentage
    for i in range(0,len(keywords)):
        for j in range(i, len(keywords)):
            # Each non-diagonal is divided by the larger of the ith and jth diagonal value to obtain percentage similarity
            if i == j:
                distance_matrix[i][j] = 100.0
            else:
                if semantic_distances[i][i] > semantic_distances[j][j]:
                    distance_matrix[i][j] = 100*semantic_distances[i][j]/semantic_distances[i][i]
                    distance_matrix[j][i] = 100*semantic_distances[j][i]/semantic_distances[i][i]
                else:
                    distance_matrix[i][j] = 100*semantic_distances[i][j]/semantic_distances[j][j]
                    distance_matrix[j][i] = 100*semantic_distances[j][i]/semantic_distances[j][j]

    return distance_matrix

def load_glove_embedding():
    # This function loads word vector data from the file glove.6B.50d.txt amd stores it in a dictionary called model
    nltk.download('stopwords')
    print("Loading GloVe Word-Vector model")
    global model
    with open("data/glove.6B.50d.txt","r", encoding = "utf-8" ) as f:
        data = f.readlines()
    for line in data:
        line = line.split()
        model[line[0]] = np.array([float(val) for val in line[1:]])
    print("Finished loading GloVe model, {} words loaded".format(len(model)))

def format_text(raw_text, formatted_dictionary):
    # Cleans raw article data and stores it in a dictionary
    words_only = re.sub("[^a-zA-Z]", " ", str(raw_text.replace("'",''))).split()
    # Removes all non-alphabet characters

    stopword_list = [w.replace("'","") for w in list(stopwords.words("english"))]
    # Loads stopwords from nltk.corpus

    cleaned_words = [w for w in words_only if w not in stopword_list]
    # Remove stopwords

    for word in cleaned_words:
        # Builds dictionary from list of words
        if word not in formatted_dictionary and len(word) > 2:
            try:
                formatted_dictionary[word] = {"frequency": 1, "vector": model[word]}
            except KeyError:
                continue
        elif len(word) > 2:
            formatted_dictionary[word]["frequency"] += 1
    return formatted_dictionary
        
def shorten_dictionary(dictionary, shortened, counter):
    for word in (sorted(dictionary.items(), key=lambda x: x[1]["frequency"], reverse = True)):
        # Sort(descending) words by frequency, highest frequency words will be stored within shortened
        shortened[word[0]] = word[1]
        counter -=1 
        if counter == 0:
            return shortened
    return shortened

def calculate_distance(a1, a2, total, avg):
    # Computes a semantic score by comparing 2 dictionaries
    for w1 in a1:
        p1 = a1[w1]
        for w2 in a2:
            # The semantic score is calculated here
            try:
                total += ((p1["frequency"]*a2[w2]["frequency"])**0.5)*((0.5*(len(w2)+len(w1)))**0.5)*(1 - sp.spatial.distance.cosine(p1["vector"], a2[w2]["vector"]))
                avg += ((p1["frequency"]*a2[w2]["frequency"])**0.5)*((0.5*(len(w2)+len(w1)))**0.5)
            except KeyError:
                continue
            except TypeError:
                break
    avg = (total/avg)
    return avg

# Below is the main logic of the first(rejected) attempt at the algorithm
"""def compute_distance(a1, a2):
    vector_1 = np.mean(calc_mean(a1, []),axis=0)
    vector_2 = np.mean(calc_mean(a2, []),axis=0)
    cosine = sp.spatial.distance.cosine(vector_1, vector_2)
    print('The two keywords are similar to',round((1-cosine)*100,2),'%')
    return
def calc_mean(a, v_list):
    # global model
    for word in a:
        try:
            v_list.append(model[word])
        except:
            continue
    return v_list"""

def write_distances(input_file, distance_matrix, size, distance_file):
    # Writes calculated semantic distances to an excel sheet
    workbook = openxl.load_workbook(input_file)
    sheet = workbook.active
    for i in range(0,size):
        # This function does not work if there are more than 25 keywords
        for j in range(0,size):
            sheet["{}{}".format(chr(66+i),2+j)] = distance_matrix[i][j]

    workbook.save(distance_file)
    workbook.close()

## Problem 4

def read_distances(distance_file):
    # Reads keywords names and % semantic similarity between them from an excel file
    keywords = read_keywords(distance_file, [], 2)
    size = len(keywords)
    workbook = openxl.load_workbook(distance_file)
    sheet = workbook.active
    distance_matrix = np.zeros((size,size), dtype=int)
    for i in range(0,size):
        for j in range(0,size):
            distance_matrix[i][j] = int(sheet["{}{}".format(chr(66+i),2+j)].value)
    return keywords, distance_matrix

def sns_distance_heatmap(keywords, distance_matrix, is_mask):
    # Generates a heatmap using the semantic distance between each word
    df = pd.DataFrame(distance_matrix, index = keywords, columns = keywords)
    sns.set_theme()
    sns.set(rc={'figure.figsize':(16.2,14.7)})
    if is_mask == True:
        mask = np.zeros_like(distance_matrix)
        mask[np.triu_indices_from(mask)] = True
        for i in range(len(keywords)):
            mask[i][i] = 0
        ax = sns.heatmap(df, cmap="YlGnBu", annot = True, fmt ="d", vmin=50, square=True, mask = mask)
    else:
        ax = sns.heatmap(df, cmap="YlGnBu", annot = True, fmt ="d", vmin=50, square=True, mask = None)
    
    ax.set_xticklabels(ax.get_xticklabels(), rotation=40, ha="right")
    ax.set_yticklabels(ax.get_xticklabels(), rotation=40, ha="right")
    ax.tick_params(labelsize=13)
    disp = ax.get_figure()
    disp.savefig("data/images/heatmap.png")
    ax.cla()
    disp.clf()

def sns_distance_bar_chart(keywords, distance_matrix):
    # Generates a Bar chart for each keyword, using semantic distances as a Y axis
    for i in range(len(keywords)):
        df = pd.DataFrame({"Keywords": keywords, "Distance": distance_matrix[i]})
        sns.set_theme()
        sns.set(rc={'figure.figsize':(8,8)})
        ax = sns.barplot(data=df, x="Keywords", y="Distance", color= "gray")
        ax.set_ylabel("Distance", fontsize=20)
        ax.set_xlabel("Keywords", fontsize=20)
        ax.set_ylim(50,100)
        ax.set_title(keywords[i],fontsize=30, pad=20)
        ax.set_xticklabels(ax.get_xticklabels(), rotation=30, ha="right")
        ax.tick_params(labelsize=12)
        disp = ax.get_figure()
        disp.savefig("data/images/bar_charts/{}.png".format(keywords[i]))
        ax.cla()
        disp.clf()

## Main

def main():
    global num_articles
    keywords_file = "keywords.xlsx"
    distance_file = "distance.xlsx"

    num_articles = 100
    # Number if articles to be read from BBC
    keywords = read_keywords(keywords_file, [], 2)
    print("{} keywords were read:\n{}".format(len(keywords),', '.join(keywords)))
    save_articles(keywords)
    # Calling this functions performs the tasks of Problem 1 and 2


    num_words = 250
    # Maximum number of words to be inputted to the semantic distance algorithm
    distance_matrix = find_keyword_distances(keywords, {}, num_words)
    # Performs tasks related to Problem 3
    write_distances(keywords_file, distance_matrix, len(keywords), distance_file)
    # Result of the algorithm is written to an excel file


    keywords, distance_matrix = read_distances(distance_file)
    # Inputs for the seaborn visualizations are obtained
    heatmap_cutoff = True
    # Decides whether the upper triangle of the heatmap generated should be cutoff
    sns_distance_heatmap(keywords, distance_matrix, heatmap_cutoff)
    # Generates a heatmap for the keywords
    print("Heatmap successfully generated, go to data/images to view the visualization")
    sns_distance_bar_chart(keywords, distance_matrix)
    print("Bar-plots successfully generated, go to data/images/bar_plots to view the visualizations")
    # Generates a bar chart for each keyword


if __name__=='__main__':
    separator = "\n\n\n{}\n{}\n{}\n\n".format(100*'#',100*'#',100*'#')
    url = "https://www.bbc.co.uk/"
    global num_articles
    if os.path.exists("data") == False:
        os.mkdir("data")
    if os.path.exists("data/images") == False:
        os.mkdir("data/images")
    if os.path.exists("data/articles") == False:
        os.mkdir("data/articles")
    if os.path.exists("data/images/bar_charts") == False:
        os.mkdir("data/images/bar_charts")
    main()   